import openpyxl
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from django.core.exceptions import ValidationError
from .models import Employee, Allowance, AllowanceType, EmployeeCategory
import re

def import_employees_from_excel(excel_file):
    """
    استيراد الموظفين والبدلات من ملف Excel
    """
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # قراءة العناوين من الصف الأول
    headers = []
    for cell in ws[1]:
        if cell.value:
            # إزالة (مطلوب) أو أي محتوى داخل أقواس
            clean_header = re.sub(r'\s*\(.*?\)', '', str(cell.value)).strip()
            headers.append(clean_header)

    imported_count = 0
    allowances_count = 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # تخطي الصفوف الفارغة
            if not any(row):
                continue

            employee_data, allowances_data = extract_employee_and_allowances_data(row, headers)

            # التحقق من وجود الموظف
            employee, created = Employee.objects.get_or_create(
                employee_number=employee_data['employee_number'],
                defaults=employee_data
            )

            if created:
                imported_count += 1
            else:
                # تحديث البيانات الموجودة
                for key, value in employee_data.items():
                    setattr(employee, key, value)
                employee.save()

            # إضافة البدلات
            if allowances_data:
                for allowance_data in allowances_data:
                    try:
                        # البحث عن نوع البدل أو إنشاؤه
                        allowance_type, _ = AllowanceType.objects.get_or_create(
                            name_arabic=allowance_data['name'],
                            defaults={
                                'name': allowance_data['name'],
                                'frequency': allowance_data.get('frequency', 'MONTHLY')
                            }
                        )

                        # إضافة أو تحديث البدل
                        allowance, allowance_created = Allowance.objects.get_or_create(
                            employee=employee,
                            allowance_type=allowance_type,
                            defaults={
                                'amount': allowance_data['amount'],
                                'type': allowance_data.get('type', 'CASH'),
                                'notes': allowance_data.get('notes', ''),
                                'is_active': True
                            }
                        )

                        if not allowance_created:
                            # تحديث البدل الموجود
                            allowance.amount = allowance_data['amount']
                            allowance.type = allowance_data.get('type', 'CASH')
                            allowance.notes = allowance_data.get('notes', '')
                            allowance.is_active = True
                            allowance.save()

                        allowances_count += 1

                    except Exception as allowance_error:
                        errors.append(f"الصف {row_num} - خطأ في البدل: {str(allowance_error)}")

        except Exception as e:
            errors.append(f"الصف {row_num}: {str(e)}")

    return {
        'imported_count': imported_count,
        'allowances_count': allowances_count,
        'errors': errors
    }


def extract_employee_and_allowances_data(row, headers):
    """
    استخراج بيانات الموظف والبدلات من صف Excel
    """
    employee_data = {}
    allowances_data = []

    # خريطة العناوين العربية والإنجليزية للموظف
    field_mapping = {
        'رقم الموظف': 'employee_number',
        'employee_number': 'employee_number',
        'الاسم': 'name',
        'name': 'name',
        'الجنسية': 'nationality',
        'nationality': 'nationality',
        'تاريخ التوظيف': 'hire_date',
        'hire_date': 'hire_date',
        'رقم الهوية': 'id_number',
        'id_number': 'id_number',
        'الفئة': 'category',
        'category': 'category',
        'الراتب الأساسي': 'basic_salary',
        'basic_salary': 'basic_salary',
        'نوع التأمين': 'insurance_type',
        'insurance_type': 'insurance_type',
        'عدد الزوجات': 'num_wives',
        'num_wives': 'num_wives',
        'عدد الأبناء': 'num_children',
        'num_children': 'num_children',
        'تكلفة الاستقدام': 'recruitment_cost',
        'recruitment_cost': 'recruitment_cost',
        'تكلفة التدريب': 'training_cost',
        'training_cost': 'training_cost',
    }

    # خريطة أعمدة البدلات (البحث عن الأعمدة التي تحتوي على "بدل")
    allowance_columns = {}
    for i, header in enumerate(headers):
        header_lower = header.lower()
        if any(keyword in header_lower for keyword in ['بدل', 'allowance', 'تعويض', 'علاوة']):
            allowance_columns[i] = header

    # استخراج بيانات الموظف
    for i, header in enumerate(headers):
        if i < len(row) and row[i] is not None:
            clean_header = re.sub(r'\s*\(.*?\)', '', header).strip()
            field_name = field_mapping.get(clean_header)
            if field_name:
                value = row[i]

                # تحويل التواريخ
                if field_name == 'hire_date':
                    if isinstance(value, datetime):
                        employee_data[field_name] = value.date()
                    elif isinstance(value, str):
                        try:
                            employee_data[field_name] = datetime.strptime(value, '%Y-%m-%d').date()
                        except:
                            try:
                                employee_data[field_name] = datetime.strptime(value, '%d/%m/%Y').date()
                            except:
                                continue

                # تحويل الأرقام
                elif field_name in ['basic_salary', 'recruitment_cost', 'training_cost']:
                    employee_data[field_name] = safe_decimal(value)

                elif field_name in ['num_wives', 'num_children']:
                    employee_data[field_name] = safe_int(value)

                # تحويل الفئة
                elif field_name == 'category':
                    try:
                        category = EmployeeCategory.objects.get(name=str(value))
                        employee_data[field_name] = category  # نمرر الكائن مباشرة
                    except EmployeeCategory.DoesNotExist:
                        employee_data[field_name] = None  # أو أي معالجة في حال لم تُوجد الفئة

                # تحويل نوع التأمين
                elif field_name == 'insurance_type':
                    insurance_mapping = {
                        'أساسي': 'BASIC',
                        'شامل': 'COMPREHENSIVE',
                        'ممتاز': 'PREMIUM',
                    }
                    employee_data[field_name] = insurance_mapping.get(str(value), str(value))

                else:
                    employee_data[field_name] = str(value).strip() if value else ''

    # استخراج بيانات البدلات
    for col_index, allowance_name in allowance_columns.items():
        if col_index < len(row) and row[col_index] is not None:
            allowance_amount = safe_decimal(row[col_index])
            if allowance_amount > 0:
                # تحديد نوع البدل وتكراره بناءً على الاسم
                frequency = 'MONTHLY'  # افتراضي
                allowance_type = 'CASH'  # افتراضي

                # تحليل اسم البدل لتحديد التكرار والنوع
                allowance_name_lower = allowance_name.lower()
                if any(keyword in allowance_name_lower for keyword in ['سنوي', 'annual', 'yearly']):
                    frequency = 'ANNUAL'
                elif any(keyword in allowance_name_lower for keyword in ['مرة', 'one_time', 'bonus']):
                    frequency = 'ONE_TIME'

                if any(keyword in allowance_name_lower for keyword in ['عيني', 'in_kind', 'benefit']):
                    allowance_type = 'IN_KIND'

                allowances_data.append({
                    'name': allowance_name,
                    'amount': allowance_amount,
                    'frequency': frequency,
                    'type': allowance_type,
                    'notes': f'مستورد من Excel - {allowance_name}'
                })

    return employee_data, allowances_data


def extract_employee_data_from_row(row, headers):
    """
    استخراج بيانات الموظف من صف Excel
    """
    data = {}

    # إنشاء قاموس من البيانات
    row_data = {}
    for i, header in enumerate(headers):
        if i < len(row):
            row_data[header] = row[i]

    # استخراج البيانات الأساسية
    try:
        data['employee_number'] = str(row_data.get('رقم الموظف', '')).strip()
        data['name'] = str(row_data.get('الاسم', '')).strip()
        data['nationality'] = str(row_data.get('الجنسية', '')).strip()

        if not all([data['employee_number'], data['name'], data['nationality']]):
            return None

        # الراتب الأساسي
        basic_salary = row_data.get('الراتب الأساسي', 0)
        if isinstance(basic_salary, (int, float)):
            data['basic_salary'] = Decimal(str(basic_salary))
        elif isinstance(basic_salary, str):
            try:
                data['basic_salary'] = Decimal(basic_salary.replace(',', ''))
            except (InvalidOperation, ValueError):
                data['basic_salary'] = Decimal('0')
        else:
            data['basic_salary'] = Decimal('0')

        # تاريخ التوظيف
        hire_date = row_data.get('تاريخ التوظيف')
        if isinstance(hire_date, datetime):
            data['hire_date'] = hire_date.date()
        elif isinstance(hire_date, date):
            data['hire_date'] = hire_date
        elif isinstance(hire_date, str):
            try:
                data['hire_date'] = datetime.strptime(hire_date, '%Y-%m-%d').date()
            except ValueError:
                try:
                    data['hire_date'] = datetime.strptime(hire_date, '%d/%m/%Y').date()
                except ValueError:
                    data['hire_date'] = date.today()
        else:
            data['hire_date'] = date.today()

        # الفئة
        category_name = str(row_data.get('الفئة', '')).strip()

        try:
            category_obj = EmployeeCategory.objects.get(name=category_name)
            data['category'] = category_obj  # لأن الحقل هو ForeignKey
        except EmployeeCategory.DoesNotExist:
            data['category'] = None  # أو يمكنك إنشاء الفئة تلقائيًا أو تجاهل السطر


        # رقم الهوية
        data['id_number'] = str(row_data.get('رقم الهوية', '')).strip()

        # نوع التأمين
        insurance_type = str(row_data.get('نوع التأمين', 'BASIC')).strip()
        insurance_mapping = {
            'أساسي': 'BASIC',
            'شامل': 'COMPREHENSIVE',
            'ممتاز': 'PREMIUM'
        }
        data['insurance_type'] = insurance_mapping.get(insurance_type, 'BASIC')

        # عدد الزوجات والأبناء
        data['num_wives'] = safe_int(row_data.get('عدد الزوجات', 0))
        data['num_children'] = safe_int(row_data.get('عدد الأبناء', 0))

        # التكاليف الإضافية
        data['recruitment_cost'] = safe_decimal(row_data.get('تكلفة الاستقدام', 0))
        data['training_cost'] = safe_decimal(row_data.get('تكلفة التدريب', 0))

        # الحالة
        data['is_active'] = True

    except Exception as e:
        print(f"خطأ في استخراج البيانات: {str(e)}")
        return None

    return data


def safe_int(value, default=0):
    """تحويل آمن إلى عدد صحيح"""
    try:
        if isinstance(value, (int, float)):
            return int(value)
        elif isinstance(value, str):
            return int(float(value.replace(',', '')))
        else:
            return default
    except (ValueError, TypeError):
        return default


def safe_decimal(value, default=0):
    """تحويل آمن إلى Decimal"""
    try:
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        elif isinstance(value, str):
            return Decimal(value.replace(',', ''))
        else:
            return Decimal(str(default))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(str(default))


def create_default_allowance_types():
    """إنشاء أنواع البدلات الافتراضية"""

    default_allowances = [
        {'name': 'housing_allowance', 'name_arabic': 'بدل السكن', 'frequency': 'MONTHLY'},
        {'name': 'transportation_allowance', 'name_arabic': 'بدل النقل', 'frequency': 'MONTHLY'},
        {'name': 'food_allowance', 'name_arabic': 'بدل الإعاشة', 'frequency': 'MONTHLY'},
        {'name': 'risk_allowance', 'name_arabic': 'بدل المخاطر', 'frequency': 'MONTHLY'},
        {'name': 'phone_allowance', 'name_arabic': 'قيمة الهاتف', 'frequency': 'MONTHLY'},
        {'name': 'tickets', 'name_arabic': 'التذاكر', 'frequency': 'ANNUAL'},
        {'name': 'medical_insurance', 'name_arabic': 'التأمين الطبي', 'frequency': 'ANNUAL'},
        {'name': 'passport_fees', 'name_arabic': 'رسوم الجوازات', 'frequency': 'ANNUAL'},
        {'name': 'work_permit_fees', 'name_arabic': 'رسوم العمل', 'frequency': 'ANNUAL'},
        {'name': 'dependent_fees', 'name_arabic': 'رسوم المرافقين', 'frequency': 'ANNUAL'},
        {'name': 'bonus', 'name_arabic': 'مكافآت', 'frequency': 'ANNUAL'},
        {'name': 'vacation_allowance', 'name_arabic': 'بدل الإجازة', 'frequency': 'ANNUAL'},
        {'name': 'training_cost', 'name_arabic': 'تكلفة التدريب', 'frequency': 'ONE_TIME'},
        {'name': 'reentry_fees', 'name_arabic': 'رسوم إعادة الدخول', 'frequency': 'ANNUAL'},
    ]

    created_count = 0

    for allowance_data in default_allowances:
        allowance_type, created = AllowanceType.objects.get_or_create(
            name=allowance_data['name'],
            defaults={
                'name_arabic': allowance_data['name_arabic'],
                'frequency': allowance_data['frequency'],
                'is_active': True
            }
        )

        if created:
            created_count += 1

    return created_count
def export_template_excel(request):
    """إنشاء قالب Excel للاستيراد"""
    from io import BytesIO
    import xlsxwriter
    from django.http import HttpResponse
    from .models import AllowanceType  # تأكد من الاستيراد الصحيح

    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})

    # تنسيق العناوين
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#4472C4',
        'font_color': 'white',
        'align': 'center',
        'valign': 'vcenter',
        'border': 1
    })

    # إنشاء ورقة العمل
    worksheet = workbook.add_worksheet('قالب الموظفين')

    # جلب أعمدة البدلات ديناميكيًا من قاعدة البيانات
    allowance_columns = list(
        AllowanceType.objects.filter(is_active=True).values_list('name_arabic', flat=True)
    )

    # العناوين المطلوبة
    headers = [
        'رقم الموظف (مطلوب)', 'الاسم (مطلوب)', 'الجنسية (مطلوب)', 'الراتب الأساسي (مطلوب)',
        'تاريخ التوظيف', 'رقم الهوية', 'الفئة', 'نوع التأمين',
        'عدد الزوجات', 'عدد الأبناء', 'تكلفة الاستقدام', 'تكلفة التدريب'
    ] + list(allowance_columns)

    # كتابة العناوين في الصف الأول
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)

    # إضافة صف مثال — يمكنك تخصيص القيم حسب الحاجة
    example_data = [
        'EMP001', 'أحمد محمد علي', 'سعودي', '8000',
        '2024-01-15', '1234567890', 'موظفين', 'أساسي',
        '1', '2', '5000', '2000'
    ] + ['1000' for _ in allowance_columns]  # قيمة افتراضية للبدلات

    for col, data in enumerate(example_data):
        worksheet.write(1, col, data)

    # تنسيق عرض الأعمدة
    worksheet.set_column(0, len(headers) - 1, 20)

    workbook.close()
    output.seek(0)

    # تجهيز رد التحميل
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="template_employees.xlsx"'
    return response
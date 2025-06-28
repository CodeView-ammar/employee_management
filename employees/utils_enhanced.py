import openpyxl
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill
from .models import Employee, AllowanceType, Allowance


def import_employees_from_excel(excel_file):
    """
    استيراد الموظفين والبدلات من ملف Excel
    """
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # قراءة العناوين من الصف الأول
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
    
    imported_count = 0
    allowances_count = 0
    errors = []
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # تخطي الصفوف الفارغة
            if not any(row):
                continue
                
            employee_data, allowances_data = extract_employee_and_allowances_data(row, headers)
            
            # التحقق من وجود الموظف
            employee, created = Employee.objects.get_or_create(
                employee_number=employee_data['employee_number'],
                defaults=employee_data
            )
            
            if created:
                imported_count += 1
            else:
                # تحديث البيانات الموجودة
                for key, value in employee_data.items():
                    setattr(employee, key, value)
                employee.save()
            
            # إضافة البدلات
            if allowances_data:
                for allowance_data in allowances_data:
                    try:
                        # البحث عن نوع البدل أو إنشاؤه
                        allowance_type, _ = AllowanceType.objects.get_or_create(
                            name_arabic=allowance_data['name'],
                            defaults={
                                'name': allowance_data['name'],
                                'frequency': allowance_data.get('frequency', 'MONTHLY')
                            }
                        )
                        
                        # إضافة أو تحديث البدل
                        allowance, allowance_created = Allowance.objects.get_or_create(
                            employee=employee,
                            allowance_type=allowance_type,
                            defaults={
                                'amount': allowance_data['amount'],
                                'type': allowance_data.get('type', 'CASH'),
                                'notes': allowance_data.get('notes', ''),
                                'is_active': True
                            }
                        )
                        
                        if not allowance_created:
                            # تحديث البدل الموجود
                            allowance.amount = allowance_data['amount']
                            allowance.type = allowance_data.get('type', 'CASH')
                            allowance.notes = allowance_data.get('notes', '')
                            allowance.is_active = True
                            allowance.save()
                        
                        allowances_count += 1
                        
                    except Exception as allowance_error:
                        errors.append(f"الصف {row_num} - خطأ في البدل: {str(allowance_error)}")
                
        except Exception as e:
            errors.append(f"الصف {row_num}: {str(e)}")
    
    return {
        'imported_count': imported_count,
        'allowances_count': allowances_count,
        'errors': errors
    }


def extract_employee_and_allowances_data(row, headers):
    """
    استخراج بيانات الموظف والبدلات من صف Excel
    """
    employee_data = {}
    allowances_data = []
    
    # خريطة العناوين العربية والإنجليزية للموظف
    field_mapping = {
        'رقم الموظف': 'employee_number',
        'employee_number': 'employee_number',
        'الاسم': 'name',
        'name': 'name',
        'الجنسية': 'nationality',
        'nationality': 'nationality',
        'تاريخ التوظيف': 'hire_date',
        'hire_date': 'hire_date',
        'رقم الهوية': 'id_number',
        'id_number': 'id_number',
        'الفئة': 'category',
        'category': 'category',
        'الراتب الأساسي': 'basic_salary',
        'basic_salary': 'basic_salary',
        'نوع التأمين': 'insurance_type',
        'insurance_type': 'insurance_type',
        'عدد الزوجات': 'num_wives',
        'num_wives': 'num_wives',
        'عدد الأبناء': 'num_children',
        'num_children': 'num_children',

    }
    
    # خريطة أعمدة البدلات (البحث عن الأعمدة التي تحتوي على "بدل")
    allowance_columns = {}
    for i, header in enumerate(headers):
        header_lower = header.lower()
        if any(keyword in header_lower for keyword in ['بدل', 'allowance', 'تعويض', 'علاوة']):
            allowance_columns[i] = header
    
    # استخراج بيانات الموظف
    for i, header in enumerate(headers):
        if i < len(row) and row[i] is not None:
            field_name = field_mapping.get(header)
            if field_name:
                value = row[i]
                
                # تحويل التواريخ
                if field_name == 'hire_date':
                    if isinstance(value, datetime):
                        employee_data[field_name] = value.date()
                    elif isinstance(value, str):
                        try:
                            employee_data[field_name] = datetime.strptime(value, '%Y-%m-%d').date()
                        except:
                            try:
                                employee_data[field_name] = datetime.strptime(value, '%d/%m/%Y').date()
                            except:
                                continue
                
                # تحويل الأرقام
                elif field_name in ['basic_salary']:
                    employee_data[field_name] = safe_decimal(value)
                    
                elif field_name in ['num_wives', 'num_children']:
                    employee_data[field_name] = safe_int(value)
                    
                # تحويل الفئة
                elif field_name == 'category':
                    try:
                        category = EmployeeCategory.objects.get(name=str(value))
                        employee_data[field_name] = category  # نمرر الكائن مباشرة
                    except EmployeeCategory.DoesNotExist:
                        employee_data[field_name] = None  # أو أي معالجة في حال لم تُوجد الفئة

                    
                # تحويل نوع التأمين
                elif field_name == 'insurance_type':
                    insurance_mapping = {
                        'أساسي': 'BASIC',
                        'شامل': 'COMPREHENSIVE',
                        'ممتاز': 'PREMIUM',
                    }
                    employee_data[field_name] = insurance_mapping.get(str(value), str(value))
                    
                else:
                    employee_data[field_name] = str(value).strip() if value else ''
    
    # استخراج بيانات البدلات
    for col_index, allowance_name in allowance_columns.items():
        if col_index < len(row) and row[col_index] is not None:
            allowance_amount = safe_decimal(row[col_index])
            if allowance_amount > 0:
                # تحديد نوع البدل وتكراره بناءً على الاسم
                frequency = 'MONTHLY'  # افتراضي
                allowance_type = 'CASH'  # افتراضي
                
                # تحليل اسم البدل لتحديد التكرار والنوع
                allowance_name_lower = allowance_name.lower()
                if any(keyword in allowance_name_lower for keyword in ['سنوي', 'annual', 'yearly']):
                    frequency = 'ANNUAL'
                elif any(keyword in allowance_name_lower for keyword in ['مرة', 'one_time', 'bonus']):
                    frequency = 'ONE_TIME'
                
                if any(keyword in allowance_name_lower for keyword in ['عيني', 'in_kind', 'benefit']):
                    allowance_type = 'IN_KIND'
                
                allowances_data.append({
                    'name': allowance_name,
                    'amount': allowance_amount,
                    'frequency': frequency,
                    'type': allowance_type,
                    'notes': f'مستورد من Excel - {allowance_name}'
                })
    
    return employee_data, allowances_data


def safe_int(value, default=0):
    """تحويل آمن إلى عدد صحيح"""
    try:
        if value is None or value == '':
            return default
        return int(float(value))
    except (ValueError, TypeError):
        return default


def safe_decimal(value, default=0):
    """تحويل آمن إلى Decimal"""
    try:
        if value is None or value == '':
            return Decimal(str(default))
        return Decimal(str(value))
    except (ValueError, TypeError, InvalidOperation):
        return Decimal(str(default))


def create_default_allowance_types():
    """إنشاء أنواع البدلات الافتراضية"""
    default_allowances = [
        ('بدل السكن', 'Housing Allowance', 'MONTHLY'),
        ('بدل المواصلات', 'Transportation Allowance', 'MONTHLY'),
        ('بدل الطعام', 'Food Allowance', 'MONTHLY'),
        ('بدل الهاتف', 'Phone Allowance', 'MONTHLY'),
        ('بدل الأطفال', 'Children Allowance', 'MONTHLY'),
        ('بدل الخطر', 'Risk Allowance', 'MONTHLY'),
        ('علاوة الأداء', 'Performance Bonus', 'ANNUAL'),
        ('مكافأة سنوية', 'Annual Bonus', 'ANNUAL'),
        ('تأمين طبي إضافي', 'Additional Medical Insurance', 'MONTHLY'),
        ('وجبات العمل', 'Work Meals', 'MONTHLY'),
    ]
    
    for name_ar, name_en, frequency in default_allowances:
        AllowanceType.objects.get_or_create(
            name_arabic=name_ar,
            defaults={
                'name': name_en,
                'frequency': frequency,
                'is_active': True
            }
        )


def export_template_excel():
    """إنشاء قالب Excel للاستيراد مع البدلات"""
    
    # إنشاء workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "قالب استيراد الموظفين والبدلات"
    
    # العناوين الأساسية للموظف
    basic_headers = [
        'رقم الموظف', 'الاسم', 'الجنسية', 'تاريخ التوظيف', 'رقم الهوية',
        'الفئة', 'الراتب الأساسي', 'نوع التأمين', 'عدد الزوجات', 'عدد الأبناء',
        'تكلفة الاستقدام', 'تكلفة التدريب'
    ]
    
    # أعمدة البدلات الشائعة
    allowance_headers = [
        'بدل السكن', 'بدل المواصلات', 'بدل الطعام', 'بدل الهاتف',
        'بدل الأطفال', 'بدل الخطر', 'علاوة الأداء', 'مكافأة سنوية'
    ]
    
    # دمج جميع العناوين
    all_headers = basic_headers + allowance_headers
    
    # كتابة العناوين
    for col, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        
        # تلوين أعمدة الموظف بلون مختلف عن أعمدة البدلات
        if col <= len(basic_headers):
            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
    
    # إضافة صف مثال
    example_employee = [
        'EMP001', 'أحمد محمد', 'سعودي', '2023-01-15', '1234567890',
        'موظفين', 5000, 'أساسي', 1, 2, 1000, 500
    ]
    
    # أمثلة للبدلات
    example_allowances = [
        1500,  # بدل السكن
        800,   # بدل المواصلات
        400,   # بدل الطعام
        200,   # بدل الهاتف
        300,   # بدل الأطفال
        '',    # بدل الخطر (فارغ)
        '',    # علاوة الأداء (فارغ)
        2000   # مكافأة سنوية
    ]
    
    example_row = example_employee + example_allowances
    
    for col, value in enumerate(example_row, 1):
        ws.cell(row=2, column=col, value=value)
    
    # إضافة تعليقات وإرشادات
    ws.cell(row=4, column=1, value="إرشادات الاستيراد:")
    ws.cell(row=5, column=1, value="1. املأ جميع الحقول الأساسية للموظف")
    ws.cell(row=6, column=1, value="2. أدخل قيم البدلات في الأعمدة المخصصة (اتركها فارغة إذا لم تكن متوفرة)")
    ws.cell(row=7, column=1, value="3. تواريخ التوظيف بصيغة YYYY-MM-DD أو DD/MM/YYYY")
    ws.cell(row=8, column=1, value="4. الفئات المتاحة: عمالة، موظفين، إداري، مهندس، فني")
    ws.cell(row=9, column=1, value="5. أنواع التأمين: أساسي، شامل، ممتاز")
    ws.cell(row=10, column=1, value="6. البدلات سيتم إنشاؤها تلقائياً حسب أسماء الأعمدة")
    
    # تنسيق الأعمدة
    for col in range(1, len(all_headers) + 1):
        column_letter = openpyxl.utils.get_column_letter(col)
        if col <= len(basic_headers):
            ws.column_dimensions[column_letter].width = 15
        else:
            ws.column_dimensions[column_letter].width = 12
    
    # إنشاء ورقة ثانية للبدلات المتاحة
    ws2 = wb.create_sheet("أنواع البدلات المتاحة")
    
    # قائمة أنواع البدلات الافتراضية
    allowance_types = [
        ('بدل السكن', 'شهري', 'نقدي'),
        ('بدل المواصلات', 'شهري', 'نقدي'),
        ('بدل الطعام', 'شهري', 'نقدي'),
        ('بدل الهاتف', 'شهري', 'نقدي'),
        ('بدل الأطفال', 'شهري', 'نقدي'),
        ('بدل الخطر', 'شهري', 'نقدي'),
        ('علاوة الأداء', 'سنوي', 'نقدي'),
        ('مكافأة سنوية', 'سنوي', 'نقدي'),
        ('تأمين طبي إضافي', 'شهري', 'عيني'),
        ('وجبات العمل', 'شهري', 'عيني'),
    ]
    
    # كتابة عناوين ورقة البدلات
    ws2.cell(row=1, column=1, value="نوع البدل").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="التكرار").font = Font(bold=True)
    ws2.cell(row=1, column=3, value="النوع").font = Font(bold=True)
    
    for row, (name, frequency, type_name) in enumerate(allowance_types, 2):
        ws2.cell(row=row, column=1, value=name)
        ws2.cell(row=row, column=2, value=frequency)
        ws2.cell(row=row, column=3, value=type_name)
    
    # تنسيق ورقة البدلات
    for col in range(1, 4):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # حفظ الملف
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="employee_import_template_with_allowances.xlsx"'
    
    wb.save(response)
    return response